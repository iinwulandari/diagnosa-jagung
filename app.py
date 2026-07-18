from datetime import datetime, date, timedelta
from flask import Flask, render_template, request, redirect, session, jsonify, flash, send_file
import os
import sqlite3
import numpy as np, io
from PIL import Image
from tensorflow.keras.applications.mobilenet_v2 import preprocess_input
import tensorflow as tf
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)

# =========================
# CONFIG
# =========================

app.secret_key = 'corncare_secret'

UPLOAD_FOLDER = 'static/uploads'
PROFILE_FOLDER = 'static/img/profile'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# BUAT FOLDER UPLOAD JIKA BELUM ADA
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# BUAT FOLDER DATABASE JIKA BELUM ADA
os.makedirs('database', exist_ok=True)

os.makedirs('static/img', exist_ok=True)

# =========================
# LOAD MODEL
# =========================

model = tf.saved_model.load("saved_model")

# AMBIL FUNCTION INFERENCE

infer = model.signatures["serving_default"]

# =========================
# NAMA KELAS
# =========================

class_names = [
    'Blight',
    'Common_Rust',
    'Gray_Leaf_Spot',
    'Healthy'
]

# =========================
# INIT DATABASE
# =========================

def get_db():
    conn = sqlite3.connect('database/jagung.db')
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS history (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        nama_file       TEXT,
        hasil_prediksi  TEXT,
        akurasi         REAL,
        tanggal         TEXT,
        waktu           TEXT,
        hari            TEXT,
        provinsi        TEXT DEFAULT "",
        kabupaten_kota  TEXT DEFAULT "",
        kecamatan       TEXT DEFAULT ""
    )''')
    for col in ['provinsi TEXT DEFAULT ""', 'kabupaten_kota TEXT DEFAULT ""', 'kecamatan TEXT DEFAULT ""']:
        try: c.execute(f"ALTER TABLE history ADD COLUMN {col}")
        except: pass
    c.execute('''CREATE TABLE IF NOT EXISTS penyakit (
        id            INTEGER PRIMARY KEY AUTOINCREMENT,
        nama_penyakit TEXT,
        gambar        TEXT DEFAULT "",
        gejala        TEXT DEFAULT "",
        penyebab      TEXT DEFAULT "",
        solusi        TEXT DEFAULT ""
    )''')
    for col in ['gambar TEXT DEFAULT ""', 'gejala TEXT DEFAULT ""', 'penyebab TEXT DEFAULT ""']:
        try: c.execute(f"ALTER TABLE penyakit ADD COLUMN {col}")
        except: pass
    c.execute('''CREATE TABLE IF NOT EXISTS admin (
        id          INTEGER PRIMARY KEY AUTOINCREMENT,
        nama        TEXT,
        username    TEXT UNIQUE,
        password    TEXT,
        foto_profil TEXT DEFAULT ""
    )''')
    try: c.execute('ALTER TABLE admin ADD COLUMN foto_profil TEXT DEFAULT ""')
    except: pass
    c.execute("SELECT COUNT(*) FROM admin")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO admin (nama,username,password,foto_profil) VALUES (?,?,?,?)",
                  ('Administrator','admin', generate_password_hash('admin123'), ''))
    conn.commit()
    conn.close()

init_db()

# ── UTILS ─────────────────────────────────────────────────
def preprocess_image(path):
    img = Image.open(path).convert('RGB').resize((224, 224))
    return np.expand_dims(preprocess_input(np.array(img)), 0).astype(np.float32)

def admin_ok():
    return 'admin' in session

def get_admin_foto():
    if 'admin_id' not in session: return ''
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT foto_profil FROM admin WHERE id=?", (session['admin_id'],))
    row = c.fetchone(); conn.close()
    return row['foto_profil'] if row and row['foto_profil'] else ''

app.jinja_env.globals['get_admin_foto'] = get_admin_foto

# =========================
# HOME
# =========================

@app.route('/')
def index():
    conn = get_db()
    c = conn.cursor()

    # Statistik utama
    c.execute("SELECT COUNT(*) FROM history")
    total = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM history WHERE hasil_prediksi != 'Healthy'")
    sakit = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM history WHERE hasil_prediksi = 'Healthy'")
    sehat = c.fetchone()[0]

    # Jumlah wilayah
    c.execute("SELECT COUNT(DISTINCT provinsi) FROM history WHERE provinsi != ''")
    n_prov = c.fetchone()[0]

    c.execute("SELECT COUNT(DISTINCT kabupaten_kota) FROM history WHERE kabupaten_kota != ''")
    n_kab = c.fetchone()[0]

    # Tren 7 hari terakhir
    c.execute("""
        SELECT tanggal, COUNT(*) j
        FROM history
        WHERE tanggal >= date('now','-6 days')
        GROUP BY tanggal
        ORDER BY tanggal
    """)
    tren = c.fetchall()

    tren_labels = [r['tanggal'] for r in tren]
    tren_data = [r['j'] for r in tren]

    # Distribusi penyakit
    c.execute("""
        SELECT hasil_prediksi, COUNT(*) j
        FROM history
        GROUP BY hasil_prediksi
        ORDER BY j DESC
    """)
    dist = c.fetchall()

    dist_labels = [r['hasil_prediksi'] for r in dist]
    dist_data = [r['j'] for r in dist]

    # Data peta
    c.execute("""
        SELECT
            provinsi,
            kabupaten_kota,
            kecamatan,
            COUNT(*) j,
            GROUP_CONCAT(DISTINCT hasil_prediksi) penyakit
        FROM history
        WHERE kabupaten_kota != ''
        GROUP BY kabupaten_kota
    """)
    peta_data = c.fetchall()

    # Katalog penyakit
    c.execute("SELECT * FROM penyakit")
    katalog = c.fetchall()

    conn.close()

    return render_template(
        'index.html',
        total_diagnosa=total,
        total_penyakit=sakit,
        total_sehat=sehat,
        n_prov=n_prov,
        n_kab=n_kab,
        tren_labels=tren_labels,
        tren_data=tren_data,
        dist_labels=dist_labels,
        dist_data=dist_data,
        peta_data=peta_data,
        katalog=katalog
    )
# =========================
# PREDICT
# =========================

@app.route('/predict', methods=['POST'])
def predict():
    if 'image' not in request.files:
        flash('Tidak ada gambar.', 'danger'); return redirect('/')
    file           = request.files['image']
    provinsi       = request.form.get('provinsi', '').strip()
    kabupaten_kota = request.form.get('kabupaten_kota', '').strip()
    kecamatan      = request.form.get('kecamatan', '').strip()
    if file.filename == '':
        flash('File belum dipilih.', 'danger'); return redirect('/')
    if not provinsi or not kabupaten_kota or not kecamatan:
        flash('Harap lengkapi Provinsi, Kabupaten/Kota, dan Kecamatan.', 'warning')
        return redirect('/#diagnosa')
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    name, ext = os.path.splitext(secure_filename(file.filename))
    filename  = f"{name}_{ts}{ext}"
    filepath  = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    pred = list(infer(tf.constant(preprocess_image(filepath))).values())[0].numpy()
    result     = class_names[np.argmax(pred)]
    confidence = round(float(np.max(pred) * 100), 2)
    now = datetime.now()
    conn = get_db()
    conn.execute('''INSERT INTO history
        (nama_file,hasil_prediksi,akurasi,tanggal,waktu,hari,provinsi,kabupaten_kota,kecamatan)
        VALUES (?,?,?,?,?,?,?,?,?)''',
        (filename, result, confidence,
         now.strftime('%Y-%m-%d'), now.strftime('%H:%M:%S'), now.strftime('%A'),
         provinsi, kabupaten_kota, kecamatan))
    conn.commit()
    c = conn.cursor()
    c.execute("SELECT * FROM penyakit WHERE LOWER(nama_penyakit) LIKE ?",
              (f"%{result.lower().replace('_',' ')}%",))
    info = c.fetchone()
    conn.close()
    return render_template('predict.html',
        image_file=filename, result=result, confidence=confidence,
        provinsi=provinsi, kabupaten_kota=kabupaten_kota,
        kecamatan=kecamatan, info_penyakit=info)

# =========================
# API PETA
# =========================

@app.route('/api/peta')
def api_peta():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT provinsi, kabupaten_kota, kecamatan,
                        COUNT(*) j, GROUP_CONCAT(DISTINCT hasil_prediksi) penyakit
                 FROM history WHERE kabupaten_kota!='' GROUP BY kabupaten_kota""")
    rows = c.fetchall(); conn.close()
    return jsonify([{
        'provinsi': r['provinsi'], 'kabupaten': r['kabupaten_kota'],
        'kecamatan': r['kecamatan'], 'jumlah': r['j'], 'penyakit': r['penyakit'] or ''
    } for r in rows])

# =========================
# AUTH
# =========================

@app.route('/login', methods=['GET', 'POST'])
@app.route('/admin/login', methods=['GET', 'POST'])
def login():
    if admin_ok(): return redirect('/admin/dashboard')
    if request.method == 'POST':
        u = request.form['username']; p = request.form['password']
        conn = get_db(); c = conn.cursor()
        c.execute("SELECT * FROM admin WHERE username=?", (u,))
        row = c.fetchone(); conn.close()
        if row and check_password_hash(row['password'], p):
            session['admin']      = True
            session['admin_id']   = row['id']
            session['admin_nama'] = row['nama']
            return redirect('/admin/dashboard')
        return render_template('auth/login.html', error='Username atau password salah!')
    return render_template('auth/login.html')

@app.route('/logout')
def logout():
    session.clear(); return redirect('/')

# =========================
# DASHBOARD ADMIN
# =========================

@app.route('/admin/dashboard')
def admin_dashboard():
    if not admin_ok(): return redirect('/login')
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM history");                                         total_scan    = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM history WHERE tanggal=?", (date.today().isoformat(),))
    scan_hari_ini = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM penyakit");                                        jml_penyakit  = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT provinsi) FROM history WHERE provinsi!=''");      jml_provinsi  = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT kabupaten_kota) FROM history WHERE kabupaten_kota!=''"); jml_kab = c.fetchone()[0]
    c.execute("SELECT COUNT(DISTINCT kecamatan) FROM history WHERE kecamatan!=''");    jml_kec       = c.fetchone()[0]
    c.execute("SELECT hasil_prediksi,COUNT(*) j FROM history GROUP BY hasil_prediksi ORDER BY j DESC LIMIT 1")
    r = c.fetchone(); penyakit_terbanyak = r['hasil_prediksi'] if r else '–'
    c.execute("""SELECT tanggal, COUNT(*) j FROM history
                 WHERE tanggal>=date('now','-6 days') GROUP BY tanggal ORDER BY tanggal""")
    rows = c.fetchall()
    g_labels = [r['tanggal'] for r in rows]; g_data = [r['j'] for r in rows]
    c.execute("SELECT hasil_prediksi, COUNT(*) j FROM history GROUP BY hasil_prediksi")
    rows = c.fetchall()
    d_labels = [r['hasil_prediksi'] for r in rows]; d_data = [r['j'] for r in rows]
    c.execute("""SELECT provinsi, kabupaten_kota, COUNT(*) j FROM history
                 WHERE kabupaten_kota!='' GROUP BY kabupaten_kota ORDER BY j DESC LIMIT 5""")
    top_wilayah = c.fetchall()
    conn.close()
    return render_template('admin/dashboard.html',
        total_scan=total_scan, scan_hari_ini=scan_hari_ini,
        jml_penyakit=jml_penyakit, jml_provinsi=jml_provinsi,
        jml_kab=jml_kab, jml_kec=jml_kec,
        penyakit_terbanyak=penyakit_terbanyak,
        g_labels=g_labels, g_data=g_data,
        d_labels=d_labels, d_data=d_data,
        top_wilayah=top_wilayah,
        admin_nama=session.get('admin_nama', 'Admin'))

# =========================
# API STATISTIK (untuk Chart.js di dashboard)
# =========================
 
@app.route('/admin/statistik')
def admin_statistik():
    if not admin_ok(): return redirect('/login')
    conn = get_db(); c = conn.cursor()
    fp     = request.args.get('filter', 'all')
    dari   = request.args.get('dari', '')
    sampai = request.args.get('sampai', '')
    prov_f = request.args.get('provinsi', '')
    conds  = []
    if   fp == '7hari':  conds.append("tanggal>=date('now','-6 days')")
    elif fp == '30hari': conds.append("tanggal>=date('now','-29 days')")
    elif fp == '6bulan': conds.append("tanggal>=date('now','-180 days')")
    elif fp == '1tahun': conds.append("tanggal>=date('now','-365 days')")
    if dari:   conds.append("tanggal>='%s'" % dari)
    if sampai: conds.append("tanggal<='%s'" % sampai)
    if prov_f: conds.append("provinsi='%s'" % prov_f)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    wc2   = list(conds) + ["kabupaten_kota!=''"]
    ww    = "WHERE " + " AND ".join(wc2)
    c.execute("SELECT COUNT(*) FROM history %s" % where);          total = c.fetchone()[0]
    c.execute("""SELECT hasil_prediksi, COUNT(*) j FROM history %s
                 GROUP BY hasil_prediksi ORDER BY j DESC""" % where)
    distribusi = c.fetchall()
    c.execute("""SELECT provinsi, kabupaten_kota, kecamatan,
                        COUNT(*) j, GROUP_CONCAT(DISTINCT hasil_prediksi) penyakit
                 FROM history %s GROUP BY kabupaten_kota ORDER BY j DESC""" % ww)
    per_wilayah = c.fetchall()
    c.execute("""SELECT tanggal, COUNT(*) j FROM history %s
                 GROUP BY tanggal ORDER BY tanggal""" % where)
    trend    = c.fetchall()
    t_labels = [r['tanggal'] for r in trend]
    t_data   = [r['j']       for r in trend]
    dl = [r['hasil_prediksi'] for r in distribusi]
    dd = [r['j']              for r in distribusi]
    c.execute("SELECT DISTINCT provinsi FROM history WHERE provinsi!='' ORDER BY provinsi")
    prov_list = [r['provinsi'] for r in c.fetchall()]
    conn.close()
    return render_template('admin/statistik.html',
        total=total, distribusi=distribusi, per_wilayah=per_wilayah,
        t_labels=t_labels, t_data=t_data,
        distribusi_labels=dl, distribusi_data=dd,
        filter_period=fp, dari=dari, sampai=sampai,
        prov_f=prov_f, prov_list=prov_list,
        admin_nama=session.get('admin_nama', 'Admin'))

# ══════════════════════════════════════════════════════════
#  ADMIN — RIWAYAT
# ══════════════════════════════════════════════════════════

@app.route('/admin/riwayat')
def admin_riwayat():
    if not admin_ok(): return redirect('/login')
    conn = get_db(); c = conn.cursor()
    search = request.args.get('search', '').strip()
    dari   = request.args.get('dari', '')
    sampai = request.args.get('sampai', '')
    prov_f = request.args.get('provinsi', '')
    kab_f  = request.args.get('kabupaten', '')
    conds  = []
    if search:  conds.append("(hasil_prediksi LIKE '%%%s%%' OR kabupaten_kota LIKE '%%%s%%' OR kecamatan LIKE '%%%s%%' OR provinsi LIKE '%%%s%%')" % (search,search,search,search))
    if dari:    conds.append("tanggal>='%s'" % dari)
    if sampai:  conds.append("tanggal<='%s'" % sampai)
    if prov_f:  conds.append("provinsi='%s'" % prov_f)
    if kab_f:   conds.append("kabupaten_kota='%s'" % kab_f)
    where = ("WHERE " + " AND ".join(conds)) if conds else ""
    c.execute("SELECT * FROM history %s ORDER BY id DESC" % where); data = c.fetchall()
    c.execute("SELECT DISTINCT provinsi FROM history WHERE provinsi!='' ORDER BY provinsi")
    prov_list = [r['provinsi'] for r in c.fetchall()]
    c.execute("SELECT DISTINCT kabupaten_kota FROM history WHERE kabupaten_kota!='' ORDER BY kabupaten_kota")
    kab_list = [r['kabupaten_kota'] for r in c.fetchall()]
    conn.close()
    return render_template('admin/riwayat.html',
        data=data, prov_list=prov_list, kab_list=kab_list,
        search=search, dari=dari, sampai=sampai, prov_f=prov_f, kab_f=kab_f,
        admin_nama=session.get('admin_nama', 'Admin'))

@app.route('/admin/riwayat/delete/<int:id>')
def riwayat_delete(id):
    if not admin_ok(): return redirect('/login')
    conn = get_db(); conn.execute('DELETE FROM history WHERE id=?', (id,))
    conn.commit(); conn.close()
    flash('Riwayat berhasil dihapus.', 'success')
    return redirect('/admin/riwayat')

# =========================
# EXPORT EXCEL
# =========================
@app.route('/admin/riwayat/export/excel')
def export_excel():
    if not admin_ok(): return redirect('/login')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        flash('Install openpyxl dulu: pip install openpyxl', 'danger')
        return redirect('/admin/riwayat')
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM history ORDER BY id DESC")
    rows = c.fetchall(); conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Riwayat Diagnosa"
    headers = ['No','Tanggal','Waktu','Hari','Provinsi','Kabupaten/Kota','Kecamatan','Hasil Diagnosa','Confidence (%)']
    hfill   = PatternFill("solid", fgColor="064E3B")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = hfill
        cell.alignment = Alignment(horizontal="center")
    for i, row in enumerate(rows, 1):
        ws.append([i, row['tanggal'], row['waktu'], row['hari'],
                   row['provinsi'] or '', row['kabupaten_kota'] or '',
                   row['kecamatan'] or '', row['hasil_prediksi'], row['akurasi']])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=f'riwayat_{date.today()}.xlsx', as_attachment=True)

# =========================
# EXPORT PDF
# =========================
@app.route('/admin/riwayat/export/pdf')
def export_pdf():
    if not admin_ok(): return redirect('/login')
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        flash('Install reportlab dulu: pip install reportlab', 'danger')
        return redirect('/admin/riwayat')
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM history ORDER BY id DESC")
    rows = c.fetchall(); conn.close()
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=20, bottomMargin=20)
    styles = getSampleStyleSheet(); elements = []
    elements.append(Paragraph("Laporan Riwayat Diagnosa CornScan AI", styles['Title']))
    elements.append(Paragraph("Dicetak: " + datetime.now().strftime('%d-%m-%Y %H:%M'), styles['Normal']))
    elements.append(Spacer(1, 12))
    data = [['No','Tanggal','Provinsi','Kabupaten/Kota','Kecamatan','Hasil','Confidence']]
    for i, row in enumerate(rows, 1):
        data.append([i, row['tanggal'], (row['provinsi'] or '')[:18],
                     (row['kabupaten_kota'] or '')[:18], (row['kecamatan'] or '')[:18],
                     row['hasil_prediksi'].replace('_',' '), f"{row['akurasi']}%"])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#064E3B')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#F0FDF4')]),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#E2E8F0')),
        ('ALIGN',(0,0),(-1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('PADDING',(0,0),(-1,-1),6),
    ]))
    elements.append(tbl); doc.build(elements); buf.seek(0)
    return send_file(buf, mimetype='application/pdf',
                     download_name=f'riwayat_{date.today()}.pdf', as_attachment=True)

# =========================
# ADMIN - PENYAKIT CRUD
# =========================

@app.route('/admin/penyakit')
def admin_penyakit():
    if not admin_ok(): return redirect('/login')
    conn = get_db(); c = conn.cursor()
    c.execute('SELECT * FROM penyakit'); data = c.fetchall(); conn.close()
    return render_template('admin/penyakit.html', data=data,
                           admin_nama=session.get('admin_nama', 'Admin'))

@app.route('/admin/penyakit/add', methods=['POST'])
def penyakit_add():
    if not admin_ok(): return redirect('/login')
    nama=request.form['nama_penyakit']; gejala=request.form.get('gejala','')
    penyebab=request.form.get('penyebab',''); solusi=request.form.get('solusi',''); gambar=''
    if 'gambar' in request.files and request.files['gambar'].filename:
        f=request.files['gambar']; fname=secure_filename(f.filename)
        f.save(os.path.join('static/img', fname)); gambar=fname
    conn=get_db()
    conn.execute('INSERT INTO penyakit (nama_penyakit,gambar,gejala,penyebab,solusi) VALUES (?,?,?,?,?)',
                 (nama,gambar,gejala,penyebab,solusi)); conn.commit(); conn.close()
    flash('Penyakit berhasil ditambahkan!','success'); return redirect('/admin/penyakit')

@app.route('/admin/penyakit/edit/<int:id>', methods=['POST'])
def penyakit_edit(id):
    if not admin_ok(): return redirect('/login')
    conn=get_db()
    conn.execute('UPDATE penyakit SET nama_penyakit=?,gejala=?,penyebab=?,solusi=? WHERE id=?',
                 (request.form['nama_penyakit'],request.form.get('gejala',''),
                  request.form.get('penyebab',''),request.form.get('solusi',''),id))
    conn.commit(); conn.close()
    flash('Penyakit berhasil diperbarui!','success'); return redirect('/admin/penyakit')

@app.route('/admin/penyakit/delete/<int:id>')
def penyakit_delete(id):
    if not admin_ok(): return redirect('/login')
    conn=get_db(); conn.execute('DELETE FROM penyakit WHERE id=?',(id,))
    conn.commit(); conn.close()
    flash('Penyakit berhasil dihapus!','success'); return redirect('/admin/penyakit')

# =========================
# ADMIN PROFILE
# =========================

@app.route('/admin/profile', methods=['GET', 'POST'])
def admin_profile():
    if not admin_ok(): return redirect('/login')
    conn = get_db(); c = conn.cursor()
    c.execute("SELECT * FROM admin WHERE id=?", (session['admin_id'],))
    admin_data = c.fetchone()
    if request.method == 'POST':
        action = request.form.get('action')
        if action == 'edit_profil':
            conn.execute("UPDATE admin SET nama=?,username=? WHERE id=?",
                         (request.form['nama'],request.form['username'],session['admin_id']))
            conn.commit(); session['admin_nama'] = request.form['nama']
            flash('Profil berhasil diperbarui!', 'success')
        elif action == 'ubah_password':
            if check_password_hash(admin_data['password'], request.form['old_password']):
                new_pw     = request.form['new_password']
                confirm_pw = request.form.get('confirm_password', '')
                if new_pw != confirm_pw:
                    flash('Konfirmasi password tidak cocok!', 'danger')
                else:
                    conn.execute("UPDATE admin SET password=? WHERE id=?",
                                 (generate_password_hash(new_pw), session['admin_id']))
                    conn.commit(); flash('Password berhasil diubah!', 'success')
            else:
                flash('Password lama salah!', 'danger')
        elif action == 'upload_foto':
            if 'foto' in request.files and request.files['foto'].filename:
                f    = request.files['foto']
                ext  = os.path.splitext(secure_filename(f.filename))[1]
                fname= f"profile_{session['admin_id']}_{datetime.now().strftime('%Y%m%d%H%M%S')}{ext}"
                f.save(os.path.join(PROFILE_FOLDER, fname))
                conn.execute("UPDATE admin SET foto_profil=? WHERE id=?", (fname, session['admin_id']))
                conn.commit(); flash('Foto profil berhasil diperbarui!', 'success')
        elif action == 'hapus_foto':
            old = admin_data['foto_profil']
            if old:
                try: os.remove(os.path.join(PROFILE_FOLDER, old))
                except: pass
            conn.execute("UPDATE admin SET foto_profil='' WHERE id=?", (session['admin_id'],))
            conn.commit(); flash('Foto profil berhasil dihapus.', 'success')
        conn.close(); return redirect('/admin/profile')
    conn.close()
    return render_template('admin/profile.html', admin_data=admin_data,
                           admin_nama=session.get('admin_nama', 'Admin'))

# =========================
# API JSON
# =========================

@app.route('/api/statistik/wilayah')
def api_wilayah():
    conn = get_db(); c = conn.cursor()
    c.execute("""SELECT provinsi,kabupaten_kota,kecamatan,
                        COUNT(*) j, GROUP_CONCAT(DISTINCT hasil_prediksi) penyakit
                 FROM history WHERE kabupaten_kota!='' GROUP BY kabupaten_kota""")
    rows = c.fetchall(); conn.close()
    return jsonify([{'provinsi':r['provinsi'],'wilayah':r['kabupaten_kota'],
                     'kecamatan':r['kecamatan'],'jumlah':r['j'],'penyakit':r['penyakit']} for r in rows])

# =========================
# RUN APP
# =========================
 
if __name__ == '__main__':
    app.run(debug=True)